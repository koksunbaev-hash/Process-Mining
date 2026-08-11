from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import UserProfile
from apps.equipment.models import MeasuringEquipment
from apps.inspections.models import InspectionCard, InspectionTask
from apps.inspections.services import complete_card, start_task
from apps.nonconformities.models import CorrectiveAction, DefectType, Nonconformity
from apps.nonconformities.services import register_nonconformity
from apps.quality.models import (
    ControlParameter,
    ControlPost,
    ControlRoute,
    ControlRouteStep,
    ControlType,
    Department,
    InspectionTemplate,
    InspectionTemplateParameter,
    QualityObject,
)
from apps.quality.services import create_quality_object_with_route
from apps.reports.services import build_report


class ReportTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_user("admin", password="x", is_staff=True, is_superuser=True)
        self.manager = User.objects.create_user("manager", password="x")
        self.manager.profile.role = UserProfile.Role.MANAGER
        self.manager.profile.save()
        self.inspector = User.objects.create_user("inspector", password="x")
        self.inspector.profile.role = UserProfile.Role.USER
        self.inspector.profile.save()
        self.master = User.objects.create_user("master", password="x")
        self.executor = User.objects.create_user("executor", password="x")
        self.blocked = User.objects.create_user("blocked", password="x")
        self.blocked.profile.role = "blocked"
        self.blocked.profile.save()

        self.department = Department.objects.create(name="Служба качества", code="QCD")
        self.other_department = Department.objects.create(name="Производство", code="PRD")
        for user in [self.manager, self.inspector, self.master, self.executor]:
            user.profile.department = self.department
            user.profile.save()

        self.control_type = ControlType.objects.create(name="измерительный", code="measure")
        self.post = ControlPost.objects.create(code="P01", name="Входной контроль", department=self.department, control_type=self.control_type, responsible_user=self.inspector)
        self.parameter = ControlParameter.objects.create(code="LEN", name="Длина", unit="мм", value_type="number", lower_limit=10, upper_limit=20)
        self.template = InspectionTemplate.objects.create(name="Карта", code="TPL", control_post=self.post, control_type=self.control_type)
        InspectionTemplateParameter.objects.create(template=self.template, parameter=self.parameter, sequence=1)
        self.route = ControlRoute.objects.create(name="Маршрут", code="R1")
        ControlRouteStep.objects.create(route=self.route, control_post=self.post, inspection_template=self.template, sequence=1)
        self.equipment = MeasuringEquipment.objects.create(name="Штангенциркуль", equipment_type="инструмент", serial_number="EQ1", department=self.department, next_verification_date=timezone.localdate() + timedelta(days=60))
        self.expired_equipment = MeasuringEquipment.objects.create(name="Манометр", equipment_type="манометр", serial_number="EQ2", department=self.department, next_verification_date=timezone.localdate() - timedelta(days=2))
        self.defect = DefectType.objects.create(code="D1", name="Трещина", criticality="critical", object_block_required=True)

        self.obj = create_quality_object_with_route(
            user=self.admin,
            unique_number="OBJ-1",
            object_type="part",
            product_name="Деталь",
            batch_number="B-1",
            quantity=1,
            department=self.department,
            route=self.route,
        )
        self.card = start_task(self.obj.tasks.first(), self.inspector)
        for result in self.card.results.all():
            result.numeric_value = 15
            result.measuring_equipment = self.equipment
            result.save()
        complete_card(self.card, self.inspector)

        self.bad_obj = create_quality_object_with_route(
            user=self.admin,
            unique_number="OBJ-2",
            object_type="part",
            product_name="Деталь с дефектом",
            batch_number="B-2",
            quantity=1,
            department=self.department,
            route=self.route,
        )
        self.bad_card = start_task(self.bad_obj.tasks.first(), self.inspector)
        self.nc = register_nonconformity(
            user=self.inspector,
            quality_object=self.bad_obj,
            inspection_card=self.bad_card,
            control_post=self.post,
            defect_type=self.defect,
            description="Критический дефект",
            criticality="critical",
            responsible_department=self.department,
            responsible_user=self.master,
            due_at=timezone.now() - timedelta(days=9),
        )
        CorrectiveAction.objects.create(nonconformity=self.nc, title="Исправить", action_plan="План", assigned_to=self.executor, due_at=timezone.now() - timedelta(days=1), status="in_progress")

    def login(self, user=None):
        self.client.force_login(user or self.manager)

    def test_index_allowed_user(self):
        self.login()
        self.assertEqual(self.client.get(reverse("reports:index")).status_code, 200)

    def test_anonymous_redirects_to_login(self):
        response = self.client.get(reverse("reports:index"))
        self.assertEqual(response.status_code, 302)

    def test_user_without_right_gets_403(self):
        self.login(self.blocked)
        self.assertEqual(self.client.get(reverse("reports:index")).status_code, 403)

    def test_period_filter_works(self):
        self.login()
        response = self.client.get(reverse("reports:detail", args=["inspection-journal"]), {"date_from": "2099-01-01"})
        self.assertContains(response, "0 записей")

    def test_status_filter_works(self):
        self.login()
        response = self.client.get(reverse("reports:detail", args=["inspection-journal"]), {"status": "completed"})
        self.assertContains(response, self.card.card_number)

    def test_department_filter_works(self):
        self.login()
        response = self.client.get(reverse("reports:detail", args=["inspection-journal"]), {"department": self.other_department.pk})
        self.assertContains(response, "0 записей")

    def test_csv_contains_filtered_data(self):
        self.login()
        response = self.client.get(reverse("reports:export", args=["inspection-journal", "csv"]), {"batch_number": "B-1"})
        text = response.content.decode("utf-8-sig")
        self.assertIn("OBJ-1", text)
        self.assertNotIn("OBJ-2", text)

    def test_csv_contains_russian_text(self):
        self.login()
        response = self.client.get(reverse("reports:export", args=["nonconformities", "csv"]))
        self.assertIn("Несоответствия", response.content.decode("utf-8-sig"))

    def test_excel_file_opens_and_has_headers(self):
        self.login()
        response = self.client.get(reverse("reports:export", args=["inspection-journal", "xlsx"]))
        workbook = load_workbook(BytesIO(response.content))
        values = [cell.value for row in workbook.active.iter_rows() for cell in row]
        self.assertIn("Номер карты", values)

    def test_pdf_content_type(self):
        self.login()
        response = self.client.get(reverse("reports:export", args=["inspection-journal", "pdf"]))
        self.assertEqual(response["Content-Type"], "application/pdf")

    def test_overdue_nonconformities_calculated(self):
        self.login()
        report = build_report("overdue-nonconformities", {}, self.manager, paginate=False)
        self.assertEqual(report["rows"][0][0], self.nc.number)

    def test_overdue_days_calculated(self):
        self.login()
        report = build_report("overdue-nonconformities", {}, self.manager, paginate=False)
        self.assertGreaterEqual(report["rows"][0][5], 9)

    def test_expired_equipment_category(self):
        self.login()
        report = build_report("equipment-verification", {}, self.manager, paginate=False)
        rows = {row[2]: row for row in report["rows"]}
        self.assertEqual(rows["EQ2"][11], "Просрочена")

    def test_first_pass_yield_calculated(self):
        self.login()
        report = build_report("first-pass-yield", {}, self.manager, paginate=False)
        summary = dict(report["summary"])
        self.assertEqual(summary["С первого раза"], 1)

    def test_department_restriction_for_master(self):
        self.master.profile.department = self.other_department
        self.master.profile.role = UserProfile.Role.USER
        self.master.profile.save()
        self.login(self.master)
        response = self.client.get(reverse("reports:detail", args=["inspection-journal"]))
        self.assertContains(response, "0 записей")

    def test_export_does_not_bypass_permissions(self):
        self.login(self.blocked)
        response = self.client.get(reverse("reports:export", args=["inspection-journal", "csv"]))
        self.assertEqual(response.status_code, 403)

    def test_empty_report_has_no_error(self):
        self.login()
        response = self.client.get(reverse("reports:detail", args=["nonconformities"]), {"date_from": "2099-01-01"})
        self.assertContains(response, "По выбранным фильтрам данные не найдены")

    def test_corrective_actions_report_opens(self):
        self.login()
        response = self.client.get(reverse("reports:detail", args=["corrective-actions"]))
        self.assertContains(response, "Исправить")
