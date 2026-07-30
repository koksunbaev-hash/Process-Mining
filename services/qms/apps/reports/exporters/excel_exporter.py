from datetime import date, datetime

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def export_excel(report):
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet.append([report["definition"].title])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([report["definition"].description])
    sheet.append(["Сформировано", timezone.localtime().strftime("%d.%m.%Y %H:%M")])
    sheet.append([])
    sheet.append(["Показатель", "Значение"])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    for label, value in report["summary"]:
        sheet.append([label, value])
    sheet.append([])
    header_row = sheet.max_row + 1
    sheet.append(report["headers"])
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EAF0FF")
        cell.alignment = Alignment(wrap_text=True)
    for row in report["rows"]:
        sheet.append([excel_value(value) for value in row])
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max(len(report['headers']), 1))}{sheet.max_row}"
    sheet.freeze_panes = f"A{header_row + 1}"
    for column_cells in sheet.columns:
        length = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{report["slug"]}_{date.today().isoformat()}.xlsx"'
    workbook.save(response)
    return response


def excel_value(value):
    if isinstance(value, datetime):
        if value.tzinfo:
            return timezone.localtime(value).replace(tzinfo=None)
        return value
    return value
